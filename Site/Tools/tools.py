from Questionnaire.models import Roles
from django.contrib.auth.hashers import make_password, check_password as check
from django.utils.timezone import utc
import datetime,requests,json
from rest_framework import serializers
from django.conf import settings

#表格转化成列表
class ExcelToDict:

    def __init__(self,path,sheet_page='抽奖名单'):
        import openpyxl
        self.path=path
        self.sheet_page=sheet_page
        self.workbook=openpyxl.load_workbook(path)

    @property
    def data(self):
        import json
        sheet = self.workbook.get_sheet_by_name(self.sheet_page)
        h, l = sheet.max_row, sheet.max_column
        lines = sheet['A1':'{}{}'.format(chr(ord('A') + l - 1), h)]
        d=[[cell.value for cell in line] for line in lines][1:]
        json_txt=json.dumps(d)
        self.save(json_txt)
        return json_txt

    def save(self,json_txt):
        with open('static/xxx.txt','w') as f:
            f.write(json_txt)

#加密密码
def set_md5_password(password):
    return make_password(password)


#对比密码
def check_password(password,md5_password):
    return check(password,md5_password)


#得到用户角色
def get_user_role(user):
    if not user:
        return ['匿名用户']
    roles=Roles.objects.filter(usersroles__user=user)
    return [role.role for role in roles]


#获得现在的时间.默认为世界协调时时间
def get_utc_now(tzinfo=utc):
    utc_now=datetime.datetime.utcnow().replace(tzinfo=tzinfo)
    return utc_now

#为request.data添加一个数据
def request_add(request,data):
    try:
        request.data._mutable=True
    except:
        pass
    request.data.update(data)
    try:
        request.data._mutable=False
    except:
        pass
    return request

def raise_error(data,error_type=1):
    if error_type==1:
        raise serializers.ValidationError(data)

class Error:
    def __init__(self,data):
        self.data=data


#发送短信验证码
class SendSMS_:
    '''
    get参数:
    account=用户账号
    ts=yyyyMMddHHmmss
    pswd=用户密码
    mobile=1234545,1323434
    msg=【签名】正式内容
    needstatus=是否需要状态报告，取值true或false
    product=订购的产品id
    resptype='json'
    https://120.27.244.164/msg/HttpBatchSendSM?account=QT-yybb&
    pswd=Net263yy&mobile=17686988582&
    msg=%E3%80%90%E7%93%A6%E5%8A%9B%E5%B7%A5%E5%8E%82%E3%80%911234&
    needstatus=True&resptype=json
    '''

    def __init__(self,mobile,data,log=False):
        account='QT-yybb'  #此家短信服务商提供的账号
        pswd='Net263yy'    #此家短信服务商提供的密码
        self.url='http://120.27.244.164/msg/HttpBatchSendSM'
        # log='瓦力工厂' if not log else log
        log='青葱道' if not log else log
        self.text='【{}】{}'.format(log,data)
        self.args={'account':account,'pswd':pswd,'mobile':mobile,
                   'msg':self.text,'needstatus':'True','resptype':'json'}
        self.get_url()

    def get_url(self):
        args_all=''
        for key in self.args:
            args='{}={}&'.format(key,self.args[key])
            args_all=args_all+args
        self.url=self.url+'?'+args_all[:-1]

    def send(self):
        #response = requests.get(self.url)
        #data=json.loads(response.text)
        #status=data['result']

        print('短信模块暂时不可用,程序即将启动模拟发送')
        info='模拟发送：向[{}]的手机号发送了[{}]'.format(self.args.get('mobile'),self.text)
        print(info)
        return True

        #status为空时返回True
        if not status:
            return True
        #status不为空时代表有错误发生.生成错误信息并返回False告诉调用者发送失败了
        else:
            error={103:'提交过快（同时时间请求验证码的用户过多）',104:'短信平台暂时不能响应请求',
                   107:'包含错误的手机号码',109:'无发送额度（请联系管理员）',110:'不在发送时间内',
                   111:'短信数量超出当月发送额度限制，请联系管理员'}
            error_info=error.get(status)
            if error_info:
                self.error_info=error_info
            else:
                self.error_info='未知错误'
            return False





#发送短信验证码
class SendSMS:
    '''
    get参数:
    account=用户账号
    ts=yyyyMMddHHmmss
    pswd=用户密码
    mobile=1234545,1323434
    msg=【签名】正式内容
    needstatus=是否需要状态报告，取值true或false
    product=订购的产品id
    resptype='json'
    https://120.27.244.164/msg/HttpBatchSendSM?account=QT-yybb&
    pswd=Net263yy&mobile=17686988582&
    msg=%E3%80%90%E7%93%A6%E5%8A%9B%E5%B7%A5%E5%8E%82%E3%80%911234&
    needstatus=True&resptype=json
    '''

    def __init__(self,mobile,data,log=False):
        self.test=settings.CODE_TEST
        # account='QT-yybb'  #此家短信服务商提供的账号
        # pswd='Net263yy'    #此家短信服务商提供的密码
        account=settings.CODE_USER
        pswd=settings.CODE_PASSWORD
        self.temp=(account,pswd)
        self.url='http://120.27.244.164/msg/HttpBatchSendSM'
        log='青葱道' if not log else log
        self.text='【{}】{}'.format(log,data)
        self.args={'account':account,'pswd':pswd,'mobile':mobile,
                   'msg':self.text,'needstatus':'True','resptype':'json'}
        self.get_url()

    def get_url(self):
        args_all=''
        for key in self.args:
            args='{}={}&'.format(key,self.args[key])
            args_all=args_all+args
        self.url=self.url+'?'+args_all[:-1]

    def send(self):
        if self.test:
            return self.test_send()
        response = requests.get(self.url)
        data=json.loads(response.text)
        status=data.get('result',400)

        #status为空时返回True
        if not status:
            return True
        #status不为空时代表有错误发生.生成错误信息并返回False告诉调用者发送失败了
        else:
            error={103:'提交过快（同时时间请求验证码的用户过多）',104:'短信平台暂时不能响应请求',
                   107:'包含错误的手机号码',109:'无发送额度（请联系管理员）',110:'不在发送时间内',
                   111:'短信数量超出当月发送额度限制，请联系管理员',400:'运营商没有返回正确的参数'}
            error_info=error.get(status)
            if error_info:
                self.error_info=error_info
            else:
                self.error_info='未知错误'
            return False

    def test_send(self):
        print('短信模块暂时不可用,程序即将启动模拟发送')
        info='{}:{}'.format(self.args.get('mobile'),self.text)
        print(info)
        return True

















